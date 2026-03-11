#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
業務ログ集計表 提出用ビュー構築スクリプト
- 名前マスタシート追加
- AllData から集計した提出用テーブル（AutoFilter付き）を構築
- 担当者・月で絞り込みやすい形に整える
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import datetime
import shutil
import os

BASE = '/mnt/c/Users/derds/Downloads/9-便利ツール-20260311T121922Z-1-001/9-便利ツール'
SRC  = f'{BASE}/業務ログ集計表202504-202603 - コピー (3)_提出用日本語版.xlsx'
DEST = f'{BASE}/業務ログ集計表202504-202603_提出用完成版.xlsx'

# 名前マスタ定義（後からシートを編集するだけで変更可能）
NAME_MASTER = [
    # 元データ名,   表示名,       並び順, 有効
    ('Osaki',      '大崎',        1,      1),
    ('Kobayashi',  '小林',        2,      1),
    ('Washizawa',  '鷲澤',        3,      1),
    ('Kawabe',     '河部',        4,      1),
    ('Ushiyama',   '牛山',        5,      1),
    ('Matsubara',  '松原',        6,      1),
    ('Oan',        'オアン',      7,      1),
    ('Iio',        '飯尾',        8,      1),
    ('Saito',      '齋藤',        9,      1),
]

# ========== スタイル定義 ==========
HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')  # 濃紺
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=11)
SUB_FILL      = PatternFill('solid', fgColor='D6E4F0')  # 薄青
MASTER_FILL   = PatternFill('solid', fgColor='E2EFDA')  # 薄緑
ALT_FILL      = PatternFill('solid', fgColor='F5F5F5')  # 交互行用
BORDER_THIN   = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

def apply_header(ws, row, cols, fill=None, font=None):
    f = fill or HEADER_FILL
    fn = font or HEADER_FONT
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = fn
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN

def apply_data_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else PatternFill('solid', fgColor='FFFFFF')
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical='center')

# ========== メイン処理 ==========
print(f"コピー元: {SRC}")
shutil.copy2(SRC, DEST)
print(f"コピー先: {DEST}")

wb = openpyxl.load_workbook(DEST)

# ---------- 1. 名前マスタシート ----------
MASTER_SHEET = '名前マスタ'
if MASTER_SHEET in wb.sheetnames:
    del wb[MASTER_SHEET]
ws_master = wb.create_sheet(MASTER_SHEET, 0)  # 先頭に追加

headers = ['元データ名', '表示名', '並び順', '有効(1=有効/0=無効)']
for c, h in enumerate(headers, 1):
    cell = ws_master.cell(1, c, h)
for c in range(1, 5):
    ws_master.cell(1, c).fill = PatternFill('solid', fgColor='375623')
    ws_master.cell(1, c).font = Font(bold=True, color='FFFFFF', size=11)
    ws_master.cell(1, c).alignment = Alignment(horizontal='center')

for r, (eng, jp, order, active) in enumerate(NAME_MASTER, 2):
    ws_master.cell(r, 1, eng)
    ws_master.cell(r, 2, jp)
    ws_master.cell(r, 3, order)
    ws_master.cell(r, 4, active)
    fill = MASTER_FILL if r % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for c in range(1, 5):
        ws_master.cell(r, c).fill = fill
        ws_master.cell(r, c).border = BORDER_THIN

ws_master.column_dimensions['A'].width = 18
ws_master.column_dimensions['B'].width = 14
ws_master.column_dimensions['C'].width = 10
ws_master.column_dimensions['D'].width = 22

# テーブルとして登録（フィルタ可能に）
tbl = Table(displayName='名前マスタ', ref=f'A1:D{1+len(NAME_MASTER)}')
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium7', showRowStripes=True)
ws_master.add_table(tbl)
print(f"[OK] 名前マスタシート作成完了 ({len(NAME_MASTER)}件)")

# ---------- 2. AllData を読み込み ----------
ws_all = wb['AllData']
rows = list(ws_all.iter_rows(values_only=True))
headers_all = rows[0]
data_rows = [r for r in rows[1:] if any(v is not None for v in r)]

df = pd.DataFrame(data_rows, columns=headers_all)
print(f"[OK] AllData 読み込み: {len(df)}行")

# 必要列の確認・数値化
for col in ['作図枚数', '作業時間', '残業時間']:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

# 月列を文字列に統一
df['月'] = df['月'].astype(str).str.strip()

# ---------- 3. 提出用シート再構築 ----------
SUBMIT_SHEET = '提出用'
if SUBMIT_SHEET in wb.sheetnames:
    del wb[SUBMIT_SHEET]
ws_sub = wb.create_sheet(SUBMIT_SHEET)

# 集計
group_cols = ['月', '担当者', '業務区分', '案件', '作業対象', '作業内容']
agg = df.groupby(group_cols, dropna=False).agg(
    作業時間=('作業時間', 'sum'),
    作図枚数=('作図枚数', 'sum'),
    残業時間=('残業時間', 'sum'),
).reset_index()

# 並び順: 月昇順 → 担当者 → 業務区分(メイン→サブ) → 案件 → 作業対象 → 作業内容
order_map = {name: i for i, (_, name, _, _) in enumerate(NAME_MASTER)}
agg['_担当者順'] = agg['担当者'].map(order_map).fillna(99)
agg['_区分順'] = agg['業務区分'].map({'メイン': 0, 'サブ': 1}).fillna(2)
agg = agg.sort_values(['月', '_担当者順', '_区分順', '案件', '作業対象', '作業内容'])
agg = agg.drop(columns=['_担当者順', '_区分順'])

print(f"[OK] 集計完了: {len(agg)}行")

# ヘッダー出力
out_cols = ['月', '担当者', '業務区分', '案件', '作業対象', '作業内容', '作業時間', '作図枚数', '残業時間']
for c, h in enumerate(out_cols, 1):
    ws_sub.cell(1, c, h)
apply_header(ws_sub, 1, len(out_cols))
ws_sub.row_dimensions[1].height = 22

# データ出力
for r_idx, (_, row) in enumerate(agg.iterrows(), 2):
    alt = (r_idx % 2 == 0)
    for c_idx, col in enumerate(out_cols, 1):
        val = row[col]
        # 数値列は小数点整合
        if col in ['作業時間', '残業時間']:
            val = round(float(val), 2) if val else 0.0
        elif col == '作図枚数':
            val = int(val) if val else 0
        ws_sub.cell(r_idx, c_idx, val)
    apply_data_row(ws_sub, r_idx, len(out_cols), alt)

# テーブル（AutoFilter）として登録
last_row = len(agg) + 1
tbl_ref = f'A1:{get_column_letter(len(out_cols))}{last_row}'
tbl2 = Table(displayName='提出用データ', ref=tbl_ref)
tbl2.tableStyleInfo = TableStyleInfo(
    name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False
)
ws_sub.add_table(tbl2)

# 列幅調整
col_widths = {
    'A': 12,   # 月
    'B': 10,   # 担当者
    'C': 10,   # 業務区分
    'D': 28,   # 案件
    'E': 22,   # 作業対象
    'F': 22,   # 作業内容
    'G': 12,   # 作業時間
    'H': 12,   # 作図枚数
    'I': 12,   # 残業時間
}
for col, w in col_widths.items():
    ws_sub.column_dimensions[col].width = w

# 先頭行固定（ヘッダーを常に見える）
ws_sub.freeze_panes = 'A2'

print(f"[OK] 提出用シート作成完了: {last_row-1}データ行")

# ---------- 4. 説明シート追加 ----------
HELP_SHEET = '使い方'
if HELP_SHEET in wb.sheetnames:
    del wb[HELP_SHEET]
ws_help = wb.create_sheet(HELP_SHEET, 1)

help_text = [
    ('【業務ログ集計表 使い方】', True),
    ('', False),
    ('■ 提出用シートの使い方', True),
    ('1. 「提出用」シートを開く', False),
    ('2. 担当者だけ見たい場合：', False),
    ('   　→ 「担当者」列のヘッダー ▼ をクリック', False),
    ('   　→ 見たい担当者にチェックを入れてOK', False),
    ('3. 月だけ見たい場合：', False),
    ('   　→ 「月」列のヘッダー ▼ をクリック', False),
    ('   　→ 見たい月にチェックを入れてOK', False),
    ('4. 複数条件を組み合わせることも可能（例：4月の河部だけ）', False),
    ('5. フィルタ解除：列ヘッダー ▼ → 「フィルタークリア」', False),
    ('', False),
    ('■ 名前マスタの変更方法', True),
    ('1. 「名前マスタ」シートを開く', False),
    ('2. 「表示名」列を編集するだけで表示名が変わる', False),
    ('3. 「有効」列を 0 にすると集計対象外にできる', False),
    ('4. ※AllData を再作成したときに反映される', False),
    ('', False),
    ('■ 列の意味', True),
    ('月       : 年月（例: 2025-04）', False),
    ('担当者   : 作業担当者名', False),
    ('業務区分 : メイン（通常業務）/ サブ（職場・後輩教育）', False),
    ('案件     : プロジェクト名', False),
    ('作業対象 : 部品・設備名など', False),
    ('作業内容 : 作図・確認・打合せ など', False),
    ('作業時間 : 合計作業時間（時間）', False),
    ('作図枚数 : 合計作図枚数', False),
    ('残業時間 : 合計残業時間（時間）', False),
]

for r, (text, bold) in enumerate(help_text, 1):
    cell = ws_help.cell(r, 1, text)
    if bold:
        cell.font = Font(bold=True, size=11, color='1F4E79')
    else:
        cell.font = Font(size=11)

ws_help.column_dimensions['A'].width = 60

# ---------- 5. シート順序を整理 ----------
# 提出用を前に出す
sheet_order = ['使い方', '提出用', '名前マスタ']
for name in reversed(sheet_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=-wb.sheetnames.index(name))

# ---------- 6. 保存 ----------
wb.save(DEST)
print(f"\n{'='*50}")
print(f"[完了] 出力ファイル: {DEST}")
print(f"{'='*50}")
print("\n■ 追加・変更シート:")
print("  - 使い方      （新規）操作説明")
print("  - 提出用      （再構築）AutoFilter付き集計テーブル")
print("  - 名前マスタ  （新規）表示名管理テーブル")
print("\n■ 使い方:")
print("  担当者絞り込み → 「担当者」列ヘッダー ▼ をクリック")
print("  月絞り込み     → 「月」列ヘッダー ▼ をクリック")
print(f"\n■ データ件数: {len(agg)}行")

months = sorted(agg['月'].unique())
persons = sorted(agg['担当者'].unique())
print(f"■ 月範囲: {months[0]} ～ {months[-1]}")
print(f"■ 担当者: {', '.join(persons)}")
