#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
各担当者のInput_*シートにダミーデータを追加し、
AllDataを再構築して提出用シートも更新するスクリプト
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import datetime
import shutil
import random

BASE = '/mnt/c/Users/derds/Downloads/9-便利ツール-20260311T121922Z-1-001/9-便利ツール'
SRC  = f'{BASE}/業務ログ集計表202504-202603_提出用完成版.xlsx'
DEST = f'{BASE}/業務ログ集計表202504-202603_提出用完成版.xlsx'

# ========== 名前マスタ（大崎→大﨑 に修正） ==========
NAME_MASTER = [
    # 元データ名,   表示名,    並び順, 有効
    ('Osaki',      '大﨑',     1,      1),
    ('Kobayashi',  '小林',     2,      1),
    ('Washizawa',  '鷲澤',     3,      1),
    ('Kawabe',     '河部',     4,      1),
    ('Ushiyama',   '牛山',     5,      1),
    ('Matsubara',  '松原',     6,      1),
    ('Oan',        'オアン',   7,      1),
    ('Iio',        '飯尾',     8,      1),
    ('Saito',      '齋藤',     9,      1),
]
# 英語名 → 日本語名 変換辞書
ENG_TO_JP = {eng: jp for eng, jp, _, _ in NAME_MASTER}

# ========== ダミーデータ定義 ==========
PROJECTS = [
    '南プラ/分離液槽', '東プラ2号', '舞洲1号', '中部水再生センター',
    '新河岸5号', 'ポータブルロックNHC-205', '大阪東部流域',
    '西部汚泥処理', '北清掃工場', '難波ポンプ場',
]
TARGETS = [
    'ダクト', 'シュート', 'ノズル', 'ステージ', 'ダイヤフラム',
    '配管', 'フランジ', '架台', '手摺', '梯子', 'ホッパー',
]
WORKS = ['作図', '図面修正', '打ち合わせ', '現地調査', '確認', '計算書作成']
SUB_PROJECTS = ['職場', '後輩教育']
SUB_TARGETS  = ['手摺', '配管', '汎用']
SUB_WORKS    = ['重量計算', '指導', 'AM:半日休暇(-3.5時間)', '環境整備']

random.seed(42)

def make_dummy_rows(name_en: str, year_months: list) -> list:
    """指定担当者のダミーデータ行リストを生成"""
    rows = []
    for ym in year_months:
        y, m = int(ym[:4]), int(ym[5:])
        # 月に10〜18日分のデータを生成
        n_days = random.randint(10, 18)
        days_used = sorted(random.sample(range(1, 28), min(n_days, 27)))
        for d in days_used:
            try:
                dt = datetime.datetime(y, m, d)
            except ValueError:
                continue
            # メイン業務
            n_main = random.randint(1, 3)
            for _ in range(n_main):
                proj  = random.choice(PROJECTS)
                tgt   = random.choice(TARGETS)
                work  = random.choice(WORKS)
                hours = round(random.choice([3.5, 4.25, 7.75, 7.0, 6.5, 5.5, 3.0]), 2)
                zumai = random.choice([0, 0, 0, 1, 2, 3])
                zangyou = round(random.choice([0, 0, 0, 1.0, 1.5, 2.0]), 2)
                rows.append((dt, proj, tgt, work, zumai, hours, zangyou))
            # サブ業務（30%確率）
            if random.random() < 0.3:
                proj  = random.choice(SUB_PROJECTS)
                tgt   = random.choice(SUB_TARGETS)
                work  = random.choice(SUB_WORKS)
                hours = round(random.choice([1.0, 2.0, 3.5, 7.75]), 2)
                rows.append((dt, proj, tgt, work, 0, hours, 0))
    return rows

# ========== 月リスト ==========
MONTHS = [f'{y}-{m:02d}' for y in [2025, 2026] for m in range(1, 13)
          if (y == 2025 and m >= 4) or (y == 2026 and m <= 3)]

# ========== スタイル定義 ==========
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
ALT_FILL    = PatternFill('solid', fgColor='F5F5F5')
BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)

def apply_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN

def apply_data_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else PatternFill('solid', fgColor='FFFFFF')
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical='center')

# ========== ファイル読み込み ==========
shutil.copy2(SRC, DEST + '.tmp')
wb = openpyxl.load_workbook(SRC)

# ========== 1. 名前マスタ更新（大崎→大﨑） ==========
ws_master = wb['名前マスタ']
# テーブル削除して再追加
for t in list(ws_master.tables.keys()):
    del ws_master.tables[t]

headers = ['元データ名', '表示名', '並び順', '有効(1=有効/0=無効)']
for c, h in enumerate(headers, 1):
    cell = ws_master.cell(1, c, h)
    cell.fill = PatternFill('solid', fgColor='375623')
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = Alignment(horizontal='center')

for r, (eng, jp, order, active) in enumerate(NAME_MASTER, 2):
    ws_master.cell(r, 1, eng)
    ws_master.cell(r, 2, jp)
    ws_master.cell(r, 3, order)
    ws_master.cell(r, 4, active)
    fill = PatternFill('solid', fgColor='E2EFDA') if r % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for c in range(1, 5):
        ws_master.cell(r, c).fill = fill
        ws_master.cell(r, c).border = BORDER_THIN

tbl = Table(displayName='名前マスタ', ref=f'A1:D{1+len(NAME_MASTER)}')
tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium7', showRowStripes=True)
ws_master.add_table(tbl)
print("[OK] 名前マスタ更新（大﨑修正）")

# ========== 2. 各 Input_* シートにダミーデータ追加 ==========
all_data_rows = []  # AllData 再構築用

for eng, jp, _, active in NAME_MASTER:
    sheet_name = f'Input_{eng}'
    if sheet_name not in wb.sheetnames:
        print(f"  スキップ（シートなし）: {sheet_name}")
        continue

    ws_in = wb[sheet_name]

    if eng == 'Kawabe':
        # 河部は既存データをそのまま使う
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            dt, proj, tgt, work = row[0], row[1], row[2], row[3]
            zumai  = row[4] or 0
            hours  = row[5] or 0
            zangyou= row[6] or 0
            if not dt or not proj:
                continue
            month = f'{dt.year}-{dt.month:02d}'
            kubun = 'サブ' if any(k in str(proj) for k in ['職場', '後輩教育']) else 'メイン'
            key   = f'{proj}|{tgt}|{work}'
            all_data_rows.append((jp, dt, month, proj, tgt, work, zumai, hours, zangyou, kubun, key))
        print(f"[OK] {sheet_name}: 既存データ取得完了")
        continue

    # ダミーデータ生成
    dummy = make_dummy_rows(eng, MONTHS)

    # シートをクリアしてヘッダー再設定
    ws_in.delete_rows(2, ws_in.max_row)
    # ヘッダーが入っているか確認（なければ追加）
    if ws_in.cell(1, 1).value != '月日':
        ws_in.cell(1, 1, '月日')
        ws_in.cell(1, 2, '案件')
        ws_in.cell(1, 3, '作業対象')
        ws_in.cell(1, 4, '作業内容')
        ws_in.cell(1, 5, '作図枚数')
        ws_in.cell(1, 6, '作業時間')
        ws_in.cell(1, 7, '残業時間')

    for r_idx, (dt, proj, tgt, work, zumai, hours, zangyou) in enumerate(dummy, 2):
        ws_in.cell(r_idx, 1, dt)
        ws_in.cell(r_idx, 2, proj)
        ws_in.cell(r_idx, 3, tgt)
        ws_in.cell(r_idx, 4, work)
        ws_in.cell(r_idx, 5, zumai)
        ws_in.cell(r_idx, 6, hours)
        ws_in.cell(r_idx, 7, zangyou)

    print(f"[OK] {sheet_name}: {len(dummy)}行のダミーデータ追加（{jp}）")

    # AllData 用に追加
    for dt, proj, tgt, work, zumai, hours, zangyou in dummy:
        month = f'{dt.year}-{dt.month:02d}'
        kubun = 'サブ' if any(k in str(proj) for k in ['職場', '後輩教育']) else 'メイン'
        key   = f'{proj}|{tgt}|{work}'
        all_data_rows.append((jp, dt, month, proj, tgt, work, zumai, hours, zangyou, kubun, key))

# ========== 3. AllData 再構築 ==========
ws_all = wb['AllData']
# 既存データクリア（ヘッダー除く）
ws_all.delete_rows(2, ws_all.max_row)

all_headers = ['担当者', '月日', '月', '案件', '作業対象', '作業内容',
               '作図枚数', '作業時間', '残業時間', '業務区分', '業務キー']

# ヘッダー確認
if ws_all.cell(1, 1).value != '担当者':
    for c, h in enumerate(all_headers, 1):
        ws_all.cell(1, c, h)

# 並び順でソート（担当者の並び順 → 月日）
order_map = {jp: order for _, jp, order, _ in NAME_MASTER}
all_data_rows.sort(key=lambda x: (order_map.get(x[0], 99), x[1]))

for r_idx, row in enumerate(all_data_rows, 2):
    for c_idx, val in enumerate(row, 1):
        ws_all.cell(r_idx, c_idx, val)

print(f"[OK] AllData 再構築: {len(all_data_rows)}行")

# ========== 4. 提出用シート再構築 ==========
df = pd.DataFrame(all_data_rows, columns=all_headers)
for col in ['作図枚数', '作業時間', '残業時間']:
    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
df['月'] = df['月'].astype(str).str.strip()

group_cols = ['月', '担当者', '業務区分', '案件', '作業対象', '作業内容']
agg = df.groupby(group_cols, dropna=False).agg(
    作業時間=('作業時間', 'sum'),
    作図枚数=('作図枚数', 'sum'),
    残業時間=('残業時間', 'sum'),
).reset_index()

order_map2 = {jp: order for _, jp, order, _ in NAME_MASTER}
agg['_担当者順'] = agg['担当者'].map(order_map2).fillna(99)
agg['_区分順']   = agg['業務区分'].map({'メイン': 0, 'サブ': 1}).fillna(2)
agg = agg.sort_values(['月', '_担当者順', '_区分順', '案件', '作業対象', '作業内容'])
agg = agg.drop(columns=['_担当者順', '_区分順'])

SUBMIT_SHEET = '提出用'
if SUBMIT_SHEET in wb.sheetnames:
    del wb[SUBMIT_SHEET]
ws_sub = wb.create_sheet(SUBMIT_SHEET)

out_cols = ['月', '担当者', '業務区分', '案件', '作業対象', '作業内容', '作業時間', '作図枚数', '残業時間']
for c, h in enumerate(out_cols, 1):
    ws_sub.cell(1, c, h)
apply_header(ws_sub, 1, len(out_cols))
ws_sub.row_dimensions[1].height = 22

for r_idx, (_, row) in enumerate(agg.iterrows(), 2):
    alt = (r_idx % 2 == 0)
    for c_idx, col in enumerate(out_cols, 1):
        val = row[col]
        if col in ['作業時間', '残業時間']:
            val = round(float(val), 2) if val else 0.0
        elif col == '作図枚数':
            val = int(val) if val else 0
        ws_sub.cell(r_idx, c_idx, val)
    apply_data_row(ws_sub, r_idx, len(out_cols), alt)

last_row = len(agg) + 1
tbl_ref = f'A1:{get_column_letter(len(out_cols))}{last_row}'
tbl2 = Table(displayName='提出用データ', ref=tbl_ref)
tbl2.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
ws_sub.add_table(tbl2)

col_widths = {'A':12,'B':10,'C':10,'D':28,'E':22,'F':22,'G':12,'H':12,'I':12}
for col, w in col_widths.items():
    ws_sub.column_dimensions[col].width = w
ws_sub.freeze_panes = 'A2'

print(f"[OK] 提出用再構築: {last_row-1}行")

# ========== 5. シート順序を整理 ==========
for name in reversed(['使い方', '提出用', '名前マスタ']):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=-idx)

# ========== 6. 保存 ==========
wb.save(DEST)
import os
os.remove(DEST + '.tmp')

print(f"\n{'='*50}")
print(f"[完了] {DEST}")
print(f"{'='*50}")

months = sorted(agg['月'].unique())
persons = agg['担当者'].unique()
person_order = {jp: order for _, jp, order, _ in NAME_MASTER}
persons_sorted = sorted(persons, key=lambda x: person_order.get(x, 99))
print(f"■ 担当者: {', '.join(persons_sorted)}")
print(f"■ 月範囲: {months[0]} ～ {months[-1]}")
print(f"■ 提出用データ行数: {last_row-1}")
